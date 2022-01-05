import openpyxl
import csv
import os

for file in os.listdir("Chapter-16/excelSpreadsheets"):

    if file.endswith("xlsx"):

        xlfile = openpyxl.load_workbook(f"Chapter-16/excelSpreadsheets/{file}")
        sheet = xlfile.active
        file = file[:-4]
        outputFile = open(f'Chapter-16/csvs/{file}_{sheet.title}.csv', 'w', newline='')
        csvwriter = csv.writer(outputFile)

        row = sheet.max_row
        column = sheet.max_column
        a = "A"

        for i in range(1,row+1):
            line = []
            for j in range(column):
                line.append(sheet[f"{chr(ord(a)+j)}{i}"].value)
            csvwriter.writerow(line)
        outputFile.close()

