import openpyxl

file = input()
wb = openpyxl.load_workbook(f"Chapter-13/{file}")
wb_inv = openpyxl.Workbook()
sheet = wb.active
sheet_inv = wb_inv.active

max_col = sheet.max_column
max_row = sheet.max_row
a = "A"

for i in range(1, max_row+1):

    for j in range(1, max_col+1):
        print("writing", f"{chr(ord(a)+j-1)}{i}",
              ":", sheet[f"{chr(ord(a)+j-1)}{i}"].value)

        sheet_inv[f"{chr(ord(a)+i-1)}{j}"] = sheet[f"{chr(ord(a)+j-1)}{i}"].value

wb_inv.save(f"Chapter-13/conv_{file}")
